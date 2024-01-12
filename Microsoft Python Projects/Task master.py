import openpyxl
import docx
import datetime
from docx import Document
from docxtpl import DocxTemplate
import jinja2
import datetime
from datetime import date

#Number of Technicians working
TECH_COUNT = 9

#Number of Tasks to be completed
TASK_COUNT = TECH_COUNT*10

#Task IDs per deadline
BL21 = [[],[],[]]
IAD20 = [[],[],[]]
BL7 = [[],[],[]]
BL6 = [[],[],[]]
BL5 = [[],[],[]]
BL3 = [[],[],[]]
BL2 = [[],[],[]]
ASH = [[],[],[]]
BL23 = [[],[],[]]
BL24 = [[],[],[]]
BL25 = [[],[],[]]
BL30 = [[],[],[]]
BL32 = [[],[],[]]
BL33 = [[],[],[]]

def priority(p, list, id):
    match p:
        case "2023":
            list[0].append(id)
        case "2025":
            list[1].append(id)

def findtasks(list, plist, tasks):
    if tasks - len(plist) > 0:
        tasks -= len(plist)
        list[2] += plist
    elif tasks - len(plist) < 0:
        while tasks > 0:
            list[2].append(plist[tasks-1])
            tasks -= 1
    elif tasks == 0:
        return 0
    return tasks

def displaytasks(DC):
    i = 0
    while len(DC[2]) > i:
        print(DC[2][i])
        i += 1

#Open Excel
workbook = openpyxl.load_workbook("GDCOAppExport.xlsx")
sheet_1 = workbook['Tasks']

##Add Deadlines, number of tasks, and task ids to their respective array, expired
for row in range(2, sheet_1.max_row + 1,):
    #Check that there is a valid task
    if (not(sheet_1.cell(row, 1).value == "InProgress" or sheet_1.cell(row, 1).value == "Created" or sheet_1.cell(row, 1).value == "Hold")) or sheet_1.cell(row, 2).value == None or sheet_1.cell(row, 11).value != None or sheet_1.cell(row, 4).value == None or (not(sheet_1.cell(row, 8).value == "AzureRMA" or sheet_1.cell(row, 8).value == "BreakFix")):
        continue

    #Get DC/TaskID and Date[0] is year
    DC = sheet_1.cell(row, 9).value
    TaskID = sheet_1.cell(row, 2).value
    Date = sheet_1.cell(row, 4).value.split("-")
    Date = Date[0]
    
    #Add Deadlines, number of tasks, and task ids to their respective array
    if DC == "BL21":
        priority(Date, BL21, TaskID)
    if DC == "IAD20":
        priority(Date, IAD20, TaskID)
    if DC == "BL7":
        priority(Date, BL7, TaskID)
    if DC == "BL6":
        priority(Date, BL6, TaskID)
    if DC == "BL5":
        priority(Date, BL5, TaskID)
    if DC == "BL3":
        priority(Date, BL3, TaskID)
    if DC == "BL2":
        priority(Date, BL2, TaskID)
    if DC == "ASH":
        priority(Date, ASH, TaskID)
    if DC == "BL23":
        priority(Date, BL23, TaskID)
    if DC == "BL24":
        priority(Date, BL24, TaskID)
    if DC == "BL25":
        priority(Date, BL25, TaskID)
    if DC == "BL30":
        priority(Date, BL30, TaskID)
    if DC == "BL32":
        priority(Date, BL32, TaskID)
    if DC == "BL33":
        priority(Date, BL33, TaskID)

TASK_COUNT = findtasks(BL21, BL21[0], TASK_COUNT)
TASK_COUNT = findtasks(IAD20, IAD20[0], TASK_COUNT)
TASK_COUNT = findtasks(BL7, BL7[0], TASK_COUNT)
TASK_COUNT = findtasks(BL6, BL6[0], TASK_COUNT)
TASK_COUNT = findtasks(BL5, BL5[0], TASK_COUNT)
TASK_COUNT = findtasks(BL3, BL3[0], TASK_COUNT)
TASK_COUNT = findtasks(BL2, BL2[0], TASK_COUNT)
TASK_COUNT = findtasks(ASH, ASH[0], TASK_COUNT)
TASK_COUNT = findtasks(BL23, BL23[0], TASK_COUNT)
TASK_COUNT = findtasks(BL24, BL24[0], TASK_COUNT)
TASK_COUNT = findtasks(BL25, BL25[0], TASK_COUNT)
TASK_COUNT = findtasks(BL30, BL30[0], TASK_COUNT)
TASK_COUNT = findtasks(BL32, BL32[0], TASK_COUNT)
TASK_COUNT = findtasks(BL33, BL33[0], TASK_COUNT)

""" TASK_COUNT = findtasks(BL21, BL21[1], TASK_COUNT)
TASK_COUNT = findtasks(IAD20, IAD20[1], TASK_COUNT)
TASK_COUNT = findtasks(BL7, BL7[1], TASK_COUNT)
TASK_COUNT = findtasks(BL6, BL6[1], TASK_COUNT)
TASK_COUNT = findtasks(BL5, BL5[1], TASK_COUNT)
TASK_COUNT = findtasks(BL3, BL3[1], TASK_COUNT)
TASK_COUNT = findtasks(BL2, BL2[1], TASK_COUNT)
TASK_COUNT = findtasks(ASH, ASH[1], TASK_COUNT)
TASK_COUNT = findtasks(BL23, BL23[1], TASK_COUNT)
TASK_COUNT = findtasks(BL24, BL24[1], TASK_COUNT)
TASK_COUNT = findtasks(BL25, BL25[1], TASK_COUNT)
TASK_COUNT = findtasks(BL30, BL30[1], TASK_COUNT)
TASK_COUNT = findtasks(BL32, BL32[1], TASK_COUNT)
TASK_COUNT = findtasks(BL33, BL33[1], TASK_COUNT) """

print("\nBL21 \nNumber of Tasks Requiring Work Tonight: ",len(BL21[2]))
displaytasks(BL21)
print("\nIAD20 \nNumber of Tasks Requiring Work Tonight: ",len(IAD20[2]))
displaytasks(IAD20)
print("\nBL7 \nNumber of Tasks Requiring Work Tonight: ",len(BL7[2]))
displaytasks(BL7)
print("\nBL6 \nNumber of Tasks Requiring Work Tonight: ",len(BL6[2]))
displaytasks(BL6)
print("\nBL5 \nNumber of Tasks Requiring Work Tonight: ",len(BL5[2]))
displaytasks(BL5)
print("\nBL3 \nNumber of Tasks Requiring Work Tonight: ",len(BL3[2]))
displaytasks(BL3)
print("\nBL2 \nNumber of Tasks Requiring Work Tonight: ",len(BL2[2]))
displaytasks(BL2)
print("\nASH \nNumber of Tasks Requiring Work Tonight: ",len(ASH[2]))
displaytasks(ASH)
print("\nBL23 \nNumber of Tasks Requiring Work Tonight: ",len(BL23[2]))
displaytasks(BL23)
print("\nBL24 \nNumber of Tasks Requiring Work Tonight: ",len(BL24[2]))
displaytasks(BL24)
print("\nBL25 \nNumber of Tasks Requiring Work Tonight: ",len(BL25[2]))
displaytasks(BL25)
print("\nBL30 \nNumber of Tasks Requiring Work Tonight: ",len(BL30[2]))
displaytasks(BL30)
print("\nBL32 \nNumber of Tasks Requiring Work Tonight: ",len(BL32[2]))
displaytasks(BL32)
print("\nBL33 \nNumber of Tasks Requiring Work Tonight: ",len(BL33[2]))
displaytasks(BL33)
print("\n")
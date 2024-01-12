import openpyxl
import docx
from docx import Document
from docxtpl import DocxTemplate
import jinja2
from datetime import date
import shutil

#Open Excel
workbook = openpyxl.load_workbook("GDCOAppExport.xlsx")
sheet_1 = workbook['Tasks']

#Inputs emails from file
Emails = []
with open('Emails.txt', 'r') as f: 
    for line in f:
        Emails = line.split(" ")

#Inputs Destination of THA from file
Destination = ""
with open('Destination.txt', 'r') as f: 
    for line in f:
        Destination = line

#Date
today = date.today()
tonight = str(today.month) +"/"+ str(today.day) +"/"+str(today.year)

#Declarations
BL21 = [[],[],[],[]]
IAD20 = [[],[],[],[]]
BL7 = [[],[],[],[]]
BL6 = [[],[],[],[]]
BL5 = [[],[],[],[]]
BL3 = [[],[],[],[]]
BL2 = [[],[],[],[]]
ASH = [[],[],[],[]]
BL23 = [[],[],[],[]]
BL24 = [[],[],[],[]]
BL25 = [[],[],[],[]]
BL30 = [[],[],[],[]]
BL32 = [[],[],[],[]]
BL33 = [[],[],[],[]]
NAMES = ["",""]
INDEXES = [0, 0, 0, 0, 0, 0]
Codes = [62772,121001,121012,61103,62780,121011,62771,63001]

#Creates THA filename
def filename(DC):
    #THA Filename
    if today.month > 9 and today.day > 9:
        filename = NAMES[1] + "_"+ DC +"_"+ str(today.year) + str(today.month) + str(today.day) + "_THA.docx"
    elif today.month < 10 and today.day < 10:
        filename = NAMES[1] + "_"+ DC +"_"+ str(today.year) + "0"+str(today.month) + "0"+str(today.day) + "_THA.docx"
    elif today.day < 10:
        filename = NAMES[1] + "_"+ DC +"_"+ str(today.year) + str(today.month) + "0"+str(today.day) + "_THA.docx"
    elif today.month < 10:
        filename = NAMES[1] + "_"+ DC +"_"+ str(today.year) + "0"+str(today.month) + str(today.day) + "_THA.docx"
    return filename

#Function used to create completed THA
def savefile(DC, taskids, locations, filename, normal):
    #Make sure list is not empty
    if len(taskids) < 1:
        return

    #Open template for THA
    if (normal == True):
        template = DocxTemplate('THA.docx')
    else:
        template = DocxTemplate('Cable_THA.docx')

    #Create context to insert into template
    context = {
        'dc': DC,
        'taskids': taskids,
        'locations': locations,
        'title': NAMES[0],
        'tonight': tonight
    }

    #Insert context into template
    template.render(context)

    #Create THA file for the day
    template.save(filename)
    shutil.move(filename, Destination)

#Remove Duplicates from List
def dups(List):
    if len(List) == 0:
        return List
    if len(List) > 0:
        return list(set(List))

#Add taskids/colos to list
def scanfile(Email):
    for row in range(2, sheet_1.max_row + 1,):
        if sheet_1.cell(row, ID).value == None or sheet_1.cell(row, ASSIGNED).value == None or sheet_1.cell(row, DCCODE).value == None or sheet_1.cell(row, COLO).value == None or not(sheet_1.cell(row, G).value == "InProgress" or sheet_1.cell(row, G).value == "Created" or sheet_1.cell(row, G).value == "Hold"):
            continue
        else:
            Assign = sheet_1.cell(row, ASSIGNED).value.split(" ")
            if len(Assign) == 3:
                Assign[2] = Assign[2].split("<")
                Assign[2] = Assign[2][1].split(">")
                Assign[2] = Assign[2][0]
                Assign.append(Assign[2].split("@"))
                Assign[3] = Assign[3][0]
            elif len(Assign) == 4:
                Assign[3] = Assign[3].split("<")
                Assign[3] = Assign[3][1].split(">")
                Assign[3] = Assign[3][0]
                Assign.append(Assign[3].split("@"))
                Assign[4] = Assign[4][0]

        if not(Assign[2] == Email or Assign[3] == Email):
            continue
        else:
            if len(Assign) == 4:
                NAMES[0] = Assign[0] + " " + Assign[1] + " - DCT"
                NAMES[1] = Assign[3]
            elif len(Assign) == 5:
                NAMES[0] = Assign[0] + " " + Assign[1] + " " + Assign[2] + " - DCT"
                NAMES[1] = Assign[4]

        if (not(Codes.__contains__(sheet_1.cell(row, FAULT).value))):
            match sheet_1.cell(row, DCCODE).value:
                case "BL21":
                    BL21[0].append(sheet_1.cell(row, ID).value)
                    BL21[1].append(sheet_1.cell(row, COLO).value)
                case "IAD20":
                    IAD20[0].append(sheet_1.cell(row, ID).value)
                    IAD20[1].append(sheet_1.cell(row, COLO).value)
                case "BL7":
                    BL7[0].append(sheet_1.cell(row, ID).value)
                    BL7[1].append(sheet_1.cell(row, COLO).value)
                case "BL6":
                    BL6[0].append(sheet_1.cell(row, ID).value)
                    BL6[1].append(sheet_1.cell(row, COLO).value)
                case "BL5":
                    BL5[0].append(sheet_1.cell(row, ID).value)
                    BL5[1].append(sheet_1.cell(row, COLO).value)
                case "BL3":
                    BL3[0].append(sheet_1.cell(row, ID).value)
                    BL3[1].append(sheet_1.cell(row, COLO).value)
                case "BL2":
                    BL2[0].append(sheet_1.cell(row, ID).value)
                    BL2[1].append(sheet_1.cell(row, COLO).value)
                case "ASH":
                    ASH[0].append(sheet_1.cell(row, ID).value)
                    ASH[1].append(sheet_1.cell(row, COLO).value)
                case "BL23":
                    BL23[0].append(sheet_1.cell(row, ID).value)
                    BL23[1].append(sheet_1.cell(row, COLO).value)
                case "BL24":
                    BL24[0].append(sheet_1.cell(row, ID).value)
                    BL24[1].append(sheet_1.cell(row, COLO).value)
                case "BL25":
                    BL25[0].append(sheet_1.cell(row, ID).value)
                    BL25[1].append(sheet_1.cell(row, COLO).value)
                case "BL30":
                    BL30[0].append(sheet_1.cell(row, ID).value)
                    BL30[1].append(sheet_1.cell(row, COLO).value)
                case "BL32":
                    BL32[0].append(sheet_1.cell(row, ID).value)
                    BL32[1].append(sheet_1.cell(row, COLO).value)
                case "BL33":
                    BL33[0].append(sheet_1.cell(row, ID).value)
                    BL33[1].append(sheet_1.cell(row, COLO).value)
        else: 
            match sheet_1.cell(row, DCCODE).value:
                case "BL21":
                    BL21[2].append(sheet_1.cell(row, ID).value)
                    BL21[3].append(sheet_1.cell(row, COLO).value)
                case "IAD20":
                    IAD20[2].append(sheet_1.cell(row, ID).value)
                    IAD20[3].append(sheet_1.cell(row, COLO).value)
                case "BL7":
                    BL7[2].append(sheet_1.cell(row, ID).value)
                    BL7[3].append(sheet_1.cell(row, COLO).value)
                case "BL6":
                    BL6[2].append(sheet_1.cell(row, ID).value)
                    BL6[3].append(sheet_1.cell(row, COLO).value)
                case "BL5":
                    BL5[2].append(sheet_1.cell(row, ID).value)
                    BL5[3].append(sheet_1.cell(row, COLO).value)
                case "BL3":
                    BL3[2].append(sheet_1.cell(row, ID).value)
                    BL3[3].append(sheet_1.cell(row, COLO).value)
                case "BL2":
                    BL2[2].append(sheet_1.cell(row, ID).value)
                    BL2[3].append(sheet_1.cell(row, COLO).value)
                case "ASH":
                    ASH[2].append(sheet_1.cell(row, ID).value)
                    ASH[3].append(sheet_1.cell(row, COLO).value)
                case "BL23":
                    BL23[2].append(sheet_1.cell(row, ID).value)
                    BL23[3].append(sheet_1.cell(row, COLO).value)
                case "BL24":
                    BL24[2].append(sheet_1.cell(row, ID).value)
                    BL24[3].append(sheet_1.cell(row, COLO).value)
                case "BL25":
                    BL25[2].append(sheet_1.cell(row, ID).value)
                    BL25[3].append(sheet_1.cell(row, COLO).value)
                case "BL30":
                    BL30[2].append(sheet_1.cell(row, ID).value)
                    BL30[3].append(sheet_1.cell(row, COLO).value)
                case "BL32":
                    BL32[2].append(sheet_1.cell(row, ID).value)
                    BL32[3].append(sheet_1.cell(row, COLO).value)
                case "BL33":
                    BL33[2].append(sheet_1.cell(row, ID).value)
                    BL33[3].append(sheet_1.cell(row, COLO).value)

#Remove duplicates from lists
def removedups():
    BL21[1] = dups(BL21[1])
    IAD20[1] = dups(IAD20[1])
    BL7[1] = dups(BL7[1])
    BL6[1] = dups(BL6[1])
    BL5[1] = dups(BL5[1])
    BL3[1] = dups(BL3[1])
    BL2[1] = dups(BL2[1])
    ASH[1] = dups(ASH[1])
    BL23[1] = dups(BL23[1])
    BL24[1] = dups(BL24[1])
    BL25[1] = dups(BL25[1])
    BL30[1] = dups(BL30[1])
    BL32[1] = dups(BL32[1])
    BL33[1] = dups(BL33[1])
    BL21[3] = dups(BL21[3])
    IAD20[3] = dups(IAD20[3])
    BL7[3] = dups(BL7[3])
    BL6[3] = dups(BL6[3])
    BL5[3] = dups(BL5[3])
    BL3[3] = dups(BL3[3])
    BL2[3] = dups(BL2[3])
    ASH[3] = dups(ASH[3])
    BL23[3] = dups(BL23[3])
    BL24[3] = dups(BL24[3])
    BL25[3] = dups(BL25[3])
    BL30[3] = dups(BL30[3])
    BL32[3] = dups(BL32[3])
    BL33[3] = dups(BL33[3])

#Save files for DC
def createfiles():
    savefile("BL21", BL21[0], BL21[1], filename("BL21"), True)
    savefile("IAD20", IAD20[0], IAD20[1], filename("IAD20"), True)
    savefile("BL7", BL7[0], BL7[1], filename("BL7"), True)
    savefile("BL6", BL6[0], BL6[1], filename("BL6"), True)
    savefile("BL5", BL5[0], BL5[1], filename("BL5"), True)
    savefile("BL3", BL3[0], BL3[1], filename("BL3"), True)
    savefile("BL2", BL2[0], BL2[1], filename("BL2"), True)
    savefile("ASH", ASH[0], ASH[1], filename("ASH"), True)
    savefile("BL23", BL23[0], BL23[1], filename("BL23"), True)
    savefile("BL24", BL24[0], BL24[1], filename("BL24"), True)
    savefile("BL25", BL25[0], BL25[1], filename("BL25"), True)
    savefile("BL30", BL30[0], BL30[1], filename("BL30"), True)
    savefile("BL32", BL32[0], BL32[1], filename("BL32"), True)
    savefile("BL33", BL33[0], BL33[1], filename("BL33"), True)
    savefile("BL21", BL21[2], BL21[3], filename("Cable_BL21"), False)
    savefile("IAD20", IAD20[2], IAD20[3], filename("Cable_IAD20"), False)
    savefile("BL7", BL7[2], BL7[3], filename("Cable_BL7"), False)
    savefile("BL6", BL6[2], BL6[3], filename("Cable_BL6"), False)
    savefile("BL5", BL5[2], BL5[3], filename("Cable_BL5"), False)
    savefile("BL3", BL3[2], BL3[3], filename("Cable_BL3"), False)
    savefile("BL2", BL2[2], BL2[3], filename("Cable_BL2"), False)
    savefile("ASH", ASH[2], ASH[3], filename("Cable_ASH"), False)
    savefile("BL23", BL23[2], BL23[3], filename("Cable_BL23"), False)
    savefile("BL24", BL24[2], BL24[3], filename("Cable_BL24"), False)
    savefile("BL25", BL25[2], BL25[3], filename("Cable_BL25"), False)
    savefile("BL30", BL30[2], BL30[3], filename("Cable_BL30"), False)
    savefile("BL32", BL32[2], BL32[3], filename("Cable_BL32"), False)
    savefile("BL33", BL33[2], BL33[3], filename("Cable_BL33"), False)

#Finds all the indexes required
def findcolumns():
    for column in range(1, sheet_1.max_column+1,):
        match sheet_1.cell(1, column).value:
            case "State":
                INDEXES[0] = column
            case "Id":
                INDEXES[1] = column
            case "Colocation":
                INDEXES[2] = column
            case "Datacenter Code":
                INDEXES[3] = column
            case "Assigned To":
                INDEXES[4] = column
            case "Fault Code":
                INDEXES[5] = column

#Looking for and assigning indexes
findcolumns()
G = INDEXES[0]
ID = INDEXES[1]
COLO = INDEXES[2]
DCCODE = INDEXES[3]
ASSIGNED = INDEXES[4]
FAULT = INDEXES[5]

#Makes files for each email in list
for Email in Emails:
    scanfile(Email)
    removedups()
    createfiles()
    BL21 = [[],[],[],[]]
    IAD20 = [[],[],[],[]]
    BL7 = [[],[],[],[]]
    BL6 = [[],[],[],[]]
    BL5 = [[],[],[],[]]
    BL3 = [[],[],[],[]]
    BL2 = [[],[],[],[]]
    ASH = [[],[],[],[]]
    BL23 = [[],[],[],[]]
    BL24 = [[],[],[],[]]
    BL25 = [[],[],[],[]]
    BL30 = [[],[],[],[]]
    BL32 = [[],[],[],[]]
    BL33 = [[],[],[],[]]
